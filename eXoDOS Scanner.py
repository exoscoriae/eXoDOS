import re
import csv
from pathlib import Path
import configparser

# ---------------------------------------------------------
# CONFIG
# ---------------------------------------------------------
def get_launchbox_root():
    # Your own default guess (user can press Enter to accept)
    default_launchbox = Path(r"F:\Emulation\Launchbox")

    while True:
        user_input = input(
            f"Enter the path to your LaunchBox folder "
            f"[{default_launchbox}]: "
        ).strip()

        # Use default if user presses Enter
        launchbox_root = Path(user_input) if user_input else default_launchbox

        # Check LaunchBox folder exists
        if not launchbox_root.exists():
            print(f"\nError: '{launchbox_root}' does not exist. Please try again.\n")
            continue

        # Check eXoDOS subfolder exists
        exodos_path = launchbox_root / "eXo" / "eXoDOS"
        if not exodos_path.exists():
            print(
                f"\nError: \eXo\eXoDOS not detected in '{launchbox_root}'.\n"
                f"Expected folder structure: Launchbox\eXo\eXoDOS\n"
            )
            continue

        # Valid path found
        print(f"\nUsing Launchbox root: {launchbox_root}\n")
        return launchbox_root

LAUNCHBOX_ROOT = get_launchbox_root()

EXODOS_BASE = LAUNCHBOX_ROOT / "eXo" / "eXoDOS"
PCJR_BASE = LAUNCHBOX_ROOT / "eXo" / "eXoPCjr"

SCRIPT_DIR = Path(__file__).resolve().parent
DOSBOX_TXT = LAUNCHBOX_ROOT / "eXo" / "util" / "dosbox.txt"
OUTPUT_CSV = SCRIPT_DIR / "conf_scan.csv"

def ask_folder_limit():
    print("\nHow many game folders do you want to scan?")
    print("Press Enter to scan ALL folders.")
    raw = input("> ").strip()

    if raw == "":
        return None  # means "all"

    if raw.isdigit():
        return int(raw)

    print("Invalid input. Scanning all folders.")
    return None

def safe_csv_value(value):
    if isinstance(value, str) and value and value[0] in ("+", "-", "="):
        return "'" + value
    return value
    
def extract_last_folder(path_str):
    p = Path(path_str)
    parts = [x for x in p.parts if x not in (".", "..", "\\", "/")]
    return parts[-1] if parts else ""
    
# ---------------------------------------------------------
# XML Parsing
# ---------------------------------------------------------
import xml.etree.ElementTree as ET

def load_launchbox_playlists(launchbox_root: Path, platform_xml: str):
    """
    Reads LaunchBox/Data/Platforms/<platform_xml> and returns:
    { game_name : "Playlist: X; Playlist: Y" }
    """
    xml_path = launchbox_root / "Data" / "Platforms" / platform_xml
    if not xml_path.exists():
        print(f"WARNING: {platform_xml} not found at {xml_path}")
        return {}

    tree = ET.parse(xml_path)
    root = tree.getroot()

    mapping = {}

    for game in root.findall("Game"):
        app_path = game.findtext("ApplicationPath", "").strip()
        series = game.findtext("Series", "").strip()

        if not app_path:
            continue

        stem = Path(app_path).stem  # e.g. "Shattered Steel (1996)"

        playlists = []
        if series:
            parts = [p.strip() for p in series.split(";")]
            for p in parts:
                if p.lower().startswith("playlist:"):
                    playlists.append(p)

        mapping[stem] = "; ".join(playlists)

    return mapping

# ---------------------------------------------------------
# LOAD dosbox.txt → { launcher_bat : fork }
# ---------------------------------------------------------
def load_dosbox_forks(path: Path):
    mapping = {}
    if not path.exists():
        return mapping

    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if ":" not in line:
                continue
            key, exe = line.split(":", 1)
            key = key.strip().strip('"').strip("'").lower().replace("\\", "/")
            if key.endswith(".bat"):
                key = key[:-4]
            exe = exe.strip()

            # Extract fork
            if "\\" in exe:
                fork = exe.split("\\", 1)[0]
            else:
                fork = "dosbox-0.74"  # vanilla

            mapping[key] = fork

    return mapping


# ---------------------------------------------------------
# FIND LAUNCHER BAT
# ---------------------------------------------------------
def find_launcher(folder: Path):
    for f in folder.iterdir():
        if f.is_file() and f.suffix.lower() == ".bat":
            name = f.stem  # filename without .bat

            # Find all (...) groups in order
            groups = re.findall(r"\(([^)]*)\)", name)
            if not groups:
                continue

            year = None
            year_index = None

            # Find the FIRST group that looks like a year
            for i, g in enumerate(groups):
                if re.fullmatch(r"[0-9xX]{4}", g):
                    year = g
                    year_index = i
                    break

            if year is None:
                continue  # no valid year found

            # Build the clean name by removing ONLY that (year) group
            parts = re.split(r"\([^)]*\)", name)
            # parts splits on ALL parentheses, so we reconstruct manually

            # Reconstruct name keeping all groups EXCEPT the year group
            rebuilt = []
            group_counter = 0
            pos = 0

            for match in re.finditer(r"\([^)]*\)", name):
                text_before = name[pos:match.start()]
                group_text = match.group(0)  # "(...)"
                pos = match.end()

                if text_before:
                    rebuilt.append(text_before)

                if group_counter != year_index:
                    rebuilt.append(group_text)

                group_counter += 1

            # Add trailing text after last group
            if pos < len(name):
                rebuilt.append(name[pos:])

            clean_name = "".join(rebuilt).strip()

            return clean_name, year, f

    return None, None, None

# ---------------------------------------------------------
# READ FIRST LINE OF CONF (fork label)
# ---------------------------------------------------------
def read_first_line(path: Path):
    try:
        with path.open("r", encoding="utf-8") as f:
            return f.readline().strip()
    except:
        return ""


# ---------------------------------------------------------
# AUTOEXEC PARSER (forced aspect + summary)
# ---------------------------------------------------------
def parse_autoexec(conf_path: Path):
    forced_aspect = ""
    mounts = []
    imgmounts = []
    mixer_cmd = ""
    boot_cmd = ""
    final_cmd = ""
    notes = []
    last_echo_index = None
    autoexec_lines = []
    meaningful_cmds = []
    autotypes = []
    path_cmds = []
    other_cmds = []
    set_cmds = []

    in_autoexec = False
    last_echo = None

    try:
        with conf_path.open("r", encoding="utf-8") as f:
            for line in f:
                raw = line.rstrip("\n")
                stripped = raw.strip()

                # Remove leading @
                if stripped.startswith("@"):
                    stripped = stripped[1:].strip()

                # Section detection
                if stripped.lower() == "[autoexec]":
                    in_autoexec = True
                    continue
                if in_autoexec and stripped.startswith("[") and stripped.endswith("]"):
                    break

                if not in_autoexec:
                    continue
                
                autoexec_lines.append(stripped)

                low = stripped.lower()

                # forced aspect
                if low.startswith("aspect"):
                    forced_aspect = stripped

                # mount
                if low.startswith("mount "):
                    mounts.append(stripped)
                    
                # autotype
                if low.startswith("autotype"):
                    autotypes.append(stripped)

                # imgmount
                if low.startswith("imgmount "):
                    imgmounts.append(stripped)
                    
                # SET / PATH handling
                if low.startswith("set "):
                    set_cmds.append(stripped)

                # detect PATH inside SET
                if low.startswith("set path=") or low.startswith("set path "):
                    path_cmds.append(stripped)
                    continue

                if low.startswith("path="):
                    path_cmds.append(stripped)
                    continue

                # mixer
                if low.startswith("mixer "):
                    mixer_cmd = stripped

                # boot
                if low.startswith("boot "):
                    boot_cmd = stripped

                # capture all echo lines except echo off
                if low.startswith("echo") and low not in ("echo off", "@echo off"):
                    notes.append(stripped)
                    last_echo_index = len(autoexec_lines) - 1

                # final command heuristic (collect all meaningful commands)
                if stripped and not low.startswith(("mount", "imgmount", "echo", "pause", "rem", "cls", "mixer", "boot")):
                    meaningful_cmds.append(stripped)
                    
                # catch-all for other autoexec lines
                if stripped and not low.startswith((
                    "mount", "imgmount", "mixer", "boot", "echo", "pause",
                    "rem", "cls", "autotype", "path", "cd ", "set"
                )):
                    other_cmds.append(stripped)
    except:
        pass
        
    if meaningful_cmds:
        final_cmd = meaningful_cmds[-1]
    else:
        final_cmd = ""

    # If a boot command exists, it IS the executed program
    if boot_cmd:
        real_final_cmd = boot_cmd
    else:
        # Otherwise fall back to meaningful command logic
        real_final_cmd = ""
        for cmd in reversed(meaningful_cmds):
            if not cmd.lower().startswith("exit"):
                real_final_cmd = cmd
                break

    # add the line after the last echo, if any
    if last_echo_index is not None:
        if last_echo_index + 1 < len(autoexec_lines):
            notes.append("AFTER_ECHO: " + autoexec_lines[last_echo_index + 1])
            
    # ---------------------------------------------------------
    # Compute internal DOSBox path at program execution
    # ---------------------------------------------------------
    current_drive = None
    drive_paths = {}

    internal_path = ""

    for line in autoexec_lines:
        low = line.lower()

        # detect drive switch (e.g., "c:")
        if len(low) == 2 and low[1] == ":" and low[0].isalpha():
            current_drive = low[0].upper()
            drive_paths.setdefault(current_drive, [])
            continue

        # mount command
        if low.startswith("mount "):
            parts = line.split()
            if len(parts) >= 3:
                drive = parts[1].replace(":", "").upper()
                host_path = parts[2]
                folder = extract_last_folder(host_path)
                drive_paths[drive] = [folder] if folder else []
            continue
    
        # cd commands
        if low.startswith("cd "):
            if current_drive is None:
                continue

            arg = line[3:].strip()

            if arg == "\\":
                drive_paths[current_drive] = []
            elif arg == "..":
                if drive_paths[current_drive]:
                    drive_paths[current_drive].pop()
            elif arg.startswith("\\"):
                drive_paths[current_drive] = [p for p in arg[1:].split("\\") if p]
            else:
                for p in arg.split("\\"):
                    if p:
                        drive_paths[current_drive].append(p)
            continue

        # stop when we reach the real executed program
        if real_final_cmd and line.strip().lower() == real_final_cmd.lower():
            break

    # Build final internal path
    if current_drive:
        internal_path = current_drive + ":\\" + "\\".join(drive_paths.get(current_drive, []))
             
    # ---------------------------------------------------------
    # Clean up other_cmds (remove drive switches, exit, blanks)
    # ---------------------------------------------------------
    cleaned = []
    for cmd in other_cmds:
        low = cmd.lower()

        # skip drive switches like "c:" or "d:"
        if len(low) == 2 and low[1] == ":" and low[0].isalpha():
            continue

        # skip exit commands
        if low == "exit" or low.startswith("exit "):
            continue

        # skip blank or comment lines
        if not cmd or low.startswith(("rem", "::", "#")):
            continue
            
        cleaned.append(cmd)

    other_cmds = cleaned
    
    if real_final_cmd:
        other_cmds = [cmd for cmd in other_cmds if cmd.lower() != real_final_cmd.lower()]

    summary = f"mounts={mounts}; imgmounts={imgmounts}; mixer={mixer_cmd}; autotype={autotypes}; path={path_cmds}; final={final_cmd}; notes={notes}"

    return forced_aspect, summary, real_final_cmd, internal_path, other_cmds, boot_cmd, set_cmds


# ---------------------------------------------------------
# GET CONF VALUE (case-insensitive)
# ---------------------------------------------------------
def get(cfg, section, key):
    try:
        return cfg[section][key]
    except:
        return ""


# ---------------------------------------------------------
# PARSE CONF FILE
# ---------------------------------------------------------
def parse_conf(conf_path: Path):

    # ---------------------------------------------------------
    # PRE-FILTER: remove [autoexec] block before ConfigParser
    # ---------------------------------------------------------
    filtered_lines = []
    in_autoexec = False

    with conf_path.open("r", encoding="utf-8") as f:
        for line in f:
            stripped = line.strip()

            # detect start of autoexec
            if stripped.lower() == "[autoexec]":
                in_autoexec = True
                continue

            # detect end of autoexec
            if in_autoexec and stripped.startswith("[") and stripped.endswith("]"):
                in_autoexec = False
                filtered_lines.append(line)  # keep the new section header
                continue

            # skip all autoexec lines
            if in_autoexec:
                continue

            filtered_lines.append(line)

    # ---------------------------------------------------------
    # NOW parse the filtered content
    # ---------------------------------------------------------
    cfg = configparser.ConfigParser(allow_no_value=True, strict=False, interpolation=None)
    cfg.read_string("".join(filtered_lines))

    # cycles / cpu_cycles
    cycles = ""
    if cfg.has_section("cpu"):
        if "cpu_cycles" in cfg["cpu"]:
            cycles = cfg["cpu"]["cpu_cycles"]
        elif "cycles" in cfg["cpu"]:
            cycles = cfg["cpu"]["cycles"]

    # vmem_delay / vmemdelay
    vmem_delay = get(cfg, "dosbox", "vmem_delay") or get(cfg, "dosbox", "vmemdelay")

    # mouse_sensitivity / sensitivity
    mouse_sens = get(cfg, "mouse", "mouse_sensitivity") or get(cfg, "sdl", "sensitivity")

    # mt32 romdir
    mt32_romdir = get(cfg, "midi", "mt32.romdir") or get(cfg, "mt32", "romdir")

    # Disney / lpt_dac
    disney = get(cfg, "speaker", "disney") or get(cfg, "speaker", "lpt_dac")

    # Innovation: sidmodel or innova
    innovation = get(cfg, "innovation", "sidmodel") or get(cfg, "innovation", "innova")

    # Printer: parallel1=printer OR printer=true
    printer = ""
    if get(cfg, "parallel", "parallel1").lower() == "printer":
        printer = "printer"
    elif get(cfg, "printer", "printer").lower() == "true":
        printer = "true"

    # forced aspect + autoexec summary
    forced_aspect, autoexec_summary, real_final_cmd, internal_path, other_cmds, boot_cmd, set_cmds = parse_autoexec(conf_path)

    return {
        "output": get(cfg, "sdl", "output"),
        "fullresolution": get(cfg, "sdl", "fullresolution"),
        "forced_aspect": forced_aspect,
        "machine": get(cfg, "dosbox", "machine"),
        "composite": get(cfg, "composite", "composite"),
        "core": get(cfg, "cpu", "core"),
        "cputype": get(cfg, "cpu", "cputype"),
        "cycles": cycles,
        "cpu_throttle": get(cfg, "cpu", "cpu_throttle"),
        "memsize": get(cfg, "dosbox", "memsize"),
        "vmemsize": get(cfg, "dosbox", "vmemsize"),
        "vmem_delay": vmem_delay,
        "dos_rate": get(cfg, "dosbox", "dos_rate"),
        "vesa_modes": get(cfg, "dosbox", "vesa_modes"),
        "vga_render_per_scanline": get(cfg, "dosbox", "vga_render_per_scanline"),
        "mouse_sensitivity": mouse_sens,

        # Sound
        "mididevice": get(cfg, "midi", "mididevice"),
        "mpu401": get(cfg, "midi", "mpu401"),
        "mt32_romdir": mt32_romdir,
        "sbtype": get(cfg, "sblaster", "sbtype"),
        "sbmixer": get(cfg, "sblaster", "sbmixer"),
        "oplmode": get(cfg, "sblaster", "oplmode"),
        "cms": get(cfg, "sblaster", "cms"),
        "oplemu": get(cfg, "sblaster", "oplemu"),
        "tandy": get(cfg, "speaker", "tandy"),
        "gus": get(cfg, "gus", "gus"),
        "pcspeaker": get(cfg, "speaker", "pcspeaker"),
        "disney": disney,
        "innovation": innovation,
        "imfc": get(cfg, "imfc", "imfc"),
        "ps1": get(cfg, "speaker", "ps1audio"),

        # Other
        "reelmagic": get(cfg, "reelmagic", "reelmagic"),
        "joysticktype": get(cfg, "joystick", "joysticktype"),
        "serial": get(cfg, "serial", "serial1"),
        "xms": get(cfg, "dos", "xms"),
        "ems": get(cfg, "dos", "ems"),
        "umb": get(cfg, "dos", "umb"),
        "pcjr_memory_config": get(cfg, "dos", "pcjr_memory_config"),
        "file_locking": get(cfg, "dos", "file_locking"),
        "ipx": get(cfg, "ipx", "ipx"),
        "printer": printer,

        # Autoexec summary
        "autoexec_summary": autoexec_summary,
        "internal_path": internal_path,
        "boot_cmd": boot_cmd,
        "real_final_cmd": real_final_cmd,
        "other_autoexec": "; ".join(other_cmds),
        "set_cmds": "; ".join(set_cmds),
    }

# ---------------------------------------------------------
# PARSE EXCEPTION.BAT
# ---------------------------------------------------------
def parse_exception_bat(game_folder: Path):
    path = game_folder / "exception.bat"
    if not path.exists():
        return {
            "exception_exists": "",
            "exception_notes": "",
            "exception_external": "",
            "exception_other": "",
            "exception_86box": "",
            "exception_scummvm": "",
            "exception_hercules": "",
            "exception_cga": "",
            "exception_composite": "",
            "exception_ega": "",
            "exception_vga": "",
            "exception_tandy": "",
            "exception_pcjr": "",
            "exception_ibm": "",
            "exception_pcspeaker": "",
            "exception_ps1": "",
            "exception_innovation": "",
            "exception_imfc": "",
            "exception_covox_disney": "",
            "exception_cms_gb": "",
            "exception_sb_adlib_opl": "",
            "exception_roland_midi_gm": "",
            "exception_mt32": "",
            "exception_sc55": "",
        }

    # Read lines
    with path.open("r", encoding="utf-8", errors="ignore") as f:
        lines = [line.strip().lstrip("@") for line in f]
        
    # Detect external Windows executables launched before/alongside DOSBox
    external = []

    for line in lines:
        stripped = line.strip()
        low = stripped.lower()

        # ignore comments
        if low.startswith(("rem", "::", "#")):
            continue

        # ignore setconsole.exe
        if "setconsole.exe" in low:
            continue

        # ---------- CASE 1: direct EXE invocation ----------
        if ".exe" in low:
            # extract the first .exe occurrence
            exe_index = low.find(".exe")
            start_index = low.rfind(" ", 0, exe_index) + 1
            exe_path = low[start_index:exe_index + 4]  # include .exe

            # ignore DOSBox itself (only if the EXE path is inside emulators\dosbox)
            if exe_path.startswith(".\\emulators\\dosbox") or exe_path.startswith(".\\emulators/dosbox"):
                continue

            external.append(stripped)
            continue

        # ---------- CASE 2: START command without .exe ----------
        if low.startswith("start "):
            parts = stripped.split()

            # skip empty title: start "" something
            idx = 1
            if len(parts) > 1 and parts[1].startswith('"'):
                idx = 2

            if len(parts) > idx:
                target = parts[idx].lower()

                # ignore DOSBox
                if target.startswith(".\\emulators\\dosbox"):
                    continue

                # detect helper without .exe
                if target.startswith(".\\") and not target.endswith(".bat"):
                    external.append(stripped)
                    continue

    # Notes (echo blocks)
    notes = []
    last_was_echo = False
    for line in lines:
        low = line.lower()
        if low.startswith("echo ") and low not in ("echo off", "@echo off"):
            if not last_was_echo and notes:
                notes.append("---")
            notes.append(line)
            last_was_echo = True
        else:
            last_was_echo = False

    # ---------------------------------------------------------
    # Catch-all other commands, but ignore any label section
    # whose contents include .\emulators\dosbox\%dosbox%
    # ---------------------------------------------------------
    other_cmds = []  # final flattened list
    sections = []    # list of (label, [commands])

    current_label = None
    current_section = []
    section_should_be_ignored = False

    def flush_section():
        if not current_section:
            return
        if not section_should_be_ignored:
            sections.append((current_label, current_section.copy()))

    for line in lines:
        stripped = line.strip()
        low = stripped.lower()

        # Detect label start
        if stripped.startswith(":"):
            flush_section()
            current_label = stripped  # keep the label
            current_section = []
            section_should_be_ignored = False
            continue

        # Ignore section if it contains a DOSBox launcher call
        if (
            "emulators\\dosbox\\" in low
            and ("dosbox.exe" in low or "%dosbox%" in low)
        ):
            section_should_be_ignored = True

        # Skip noise lines
        if not stripped:
            continue
        if low.startswith(("rem", "::", "#", "echo", "pause", "cls")):
            continue
        if len(low) == 2 and low[1] == ":" and low[0].isalpha():
            continue
        if low.startswith("exit"):
            continue
        if "setconsole.exe" in low:
            continue
        if low.startswith("cd ") and "%" in low:
            continue

        # SET lines must always be captured
        if low.startswith("set "):
            current_section.append(stripped)
            continue

        # Otherwise add normally
        current_section.append(stripped)

    # flush last section
    flush_section()

    # Build final output with prefixes and labels
    flat = []
    for label, cmds in sections:
        flat.append("---")
        if label:
            flat.append(label)
        flat.extend(cmds)

    other_cmds = flat

    # Keyword detection
    text = "\n".join(lines).lower()
    
    # Sound Blaster / Adlib / OPL — smarter SB detection
    def sb_flag():
        # full-word or word-start SB, but NOT dosbox or sblaster
        import re
        if re.search(r"\bsb(?!laster)\b", text):   # matches "sb", "sb16", "sbpro", etc.
            return "yes"
        if "soundblaster" in text or "sound blaster" in text or "sound-blaster" in text:
            return "yes"
        if "adlib" in text:
            return "yes"
        if "opl" in text:
            return "yes"
        return ""

    def flag(*keywords):
        return "yes" if any(k in text for k in keywords) else ""

    # IBM only when used as a selectable option
    ibm_flag = ""
    if "ibm" in text:
        if any(v in text for v in ("cga", "ega", "vga", "mcga", "tandy", "pcjr", "machine")):
            ibm_flag = "yes"

    return {
        "exception_exists": "yes",
        "exception_notes": "; ".join(notes),
        "exception_external": "; ".join(external),
        "exception_other": "; ".join(other_cmds),
        "exception_86box": flag("86box"),
        "exception_scummvm": flag("scummvm"),
        "exception_hercules": flag("hercules"),
        "exception_cga": flag("cga"),
        "exception_composite": flag("composite"),
        "exception_ega": flag("ega"),
        "exception_vga": flag("vga"),
        "exception_tandy": flag("tandy"),
        "exception_pcjr": flag("pcjr"),
        "exception_ibm": ibm_flag,
        "exception_pcspeaker": flag("pc speaker", "pcspeaker", "pc-speaker"),
        "exception_ps1": flag("ps1", "ps\\1"),
        "exception_innovation": flag("innovation"),
        "exception_imfc": flag("imfc", "music feature"),
        "exception_covox_disney": flag("covox", "disney"),
        "exception_cms_gb": flag("cms", "game blaster", "gameblaster", "game-blaster", "gb "),
        "exception_sb_adlib_opl": sb_flag(),
        "exception_roland_midi_gm": flag("roland", "midi", "general midi"),
        "exception_mt32": flag("mt32", "mt-32", "mt 32", "lapc"),
        "exception_sc55": flag("soundcanvas", "sound canvas", "sound-canvas", "ssc", "sc55", "sc-55", "fluidsynth"),
    }

# ---------------------------------------------------------
# PARSE RUN.BAT CHAIN (from ZIP)
# ---------------------------------------------------------
import zipfile
import threading, time, sys

ticker_running = True
ticker_paused = False

def ticker():
    spinner = ["|", "/", "-", "\\"]
    idx = 0
    while ticker_running:
        if not ticker_paused:
            sys.stdout.write("\rScanning... " + spinner[idx])
            sys.stdout.flush()
            idx = (idx + 1) % len(spinner)
        time.sleep(0.1)
    # clear spinner when done
    sys.stdout.write("\rScanning... done.\n")
    sys.stdout.flush()

def parse_run_chain(game_folder: Path, executed_program: str, internal_path: str, zip_name: str, project_base: Path):
    debug_prefix = f"[RUN-CHAIN] {game_folder.name}: "
    """
    Returns a list of dicts:
    [
        { run1_* columns },
        { run2_* columns },
        ...
    ]
    """

    # 1. Locate the ZIP file
    if not zip_name:
        print(f"[ERROR] No launcher BAT found for game: {game_folder.name}")
        print("Cannot determine ZIP name for run.bat chain.\n")
        return []

    zip_path = project_base / zip_name
    if not zip_path.exists():
        print(f"[ERROR] Missing ZIP for game: {game_folder.name}")
        print(f"Expected ZIP at: {zip_path}")
        print("This game uses run.bat logic, but the ZIP is missing.\n")
        return []
        
    # -----------------------------------------------------
    # 2. Load all run*.bat files from ZIP
    # -----------------------------------------------------
    run_files = {}
    with zipfile.ZipFile(zip_path, "r") as z:
        for name in z.namelist():
            p = Path(name)
            if p.suffix.lower() == ".bat" and "run" in p.stem.lower():
                # store using normalized forward-slash path
                run_files[str(p).lower()] = name

    if not run_files:
        return []

    # -----------------------------------------------------
    # 3. Determine starting run file
    # -----------------------------------------------------
    if not executed_program:
        return []

    # Extract first token
    token = executed_program.strip().split()[0]
    token = token.strip('"').strip("'")

    # Remove .bat if present
    token_stem = Path(token).stem.lower()

    # Build expected DOS path inside ZIP
    # internal_path example: "C:\GAMES\DOS"
    # We only care about the folder structure after the drive letter
    zip_expected_path = None

    if internal_path:
        # Extract path after "X:\"
        parts = internal_path.split(":", 1)
        if len(parts) == 2:
            subpath = parts[1].lstrip("\\/")
            # Combine with executed program
            zip_expected_path = (Path(subpath) / (token_stem + ".bat")).as_posix().lower()

    # -----------------------------------------------------
    # 4. Try to match ZIP entry by FULL PATH
    # -----------------------------------------------------
    start_file = None

    if zip_expected_path:
        for fullpath in run_files:
            if fullpath.endswith(zip_expected_path):
                start_file = run_files[fullpath]
                break

    # -----------------------------------------------------
    # 5. If not found, try matching by stem (run, runH, run2)
    # -----------------------------------------------------
    if not start_file:
        for fullpath in run_files:
            if Path(fullpath).stem.lower() == token_stem:
                start_file = run_files[fullpath]
                break

    # -----------------------------------------------------
    # 6. Fallback to "run.bat"
    # -----------------------------------------------------
    if not start_file:
        for fullpath in run_files:
            if Path(fullpath).stem.lower() == "run":
                start_file = run_files[fullpath]
                break

    if not start_file:
        return []

    # -----------------------------------------------------
    # 4. Helper: parse a single run*.bat file
    # -----------------------------------------------------
    def parse_single_run(lines):
    
        notes = []
        last_was_echo = False

        other_cmds = []
        config_cmds = []
        
        mixer_cmds = []
        autotype_cmds = []

        sections = []  # list of (label, [commands])
        current_label = None
        current_section = []
        section_should_be_ignored = False

        def flush_section():
            if not current_section:
                return
            if not section_should_be_ignored:
                sections.append((current_label, current_section.copy()))

        for line in lines:
            stripped = line.strip().lstrip("@")
            low = stripped.lower()

            # Echo notes (unchanged)
            if low.startswith("echo ") and low not in ("echo off", "@echo off"):
                if not last_was_echo and notes:
                    notes.append("---")
                notes.append(stripped)
                last_was_echo = True
                continue
            else:
                last_was_echo = False

            # Detect label start
            if stripped.startswith(":"):
                flush_section()
                current_label = stripped
                current_section = []
                section_should_be_ignored = False
                continue

            # Ignore DOSBox launcher sections
            if (
                "emulators\\dosbox\\" in low
                and ("dosbox.exe" in low or "%dosbox%" in low)
            ):
                section_should_be_ignored = True

            # Skip noise
            if not stripped:
                continue
            if low.startswith(("rem", "::", "#", "echo", "pause", "cls")):
                continue
            if len(low) == 2 and low[1] == ":" and low[0].isalpha():
                continue
            if low.startswith("exit"):
                continue
            if "setconsole.exe" in low:
                continue
            if low.startswith("cd ") and "%" in low:
                continue

            # SET lines always included
            if low.startswith("set "):
                current_section.append(stripped)
                continue

            # config lines
            if low.startswith("config "):
                config_cmds.append(stripped)
            
            # mixer lines
            if low.startswith("mixer "):
                mixer_cmds.append(stripped)

            # autotype lines
            if low.startswith("autotype "):
                autotype_cmds.append(stripped)

            current_section.append(stripped)

        # flush last section
        flush_section()

        # Build final output with prefixes and labels
        flat = []
        for label, cmds in sections:
            flat.append("---")
            if label:
                flat.append(label)
            flat.extend(cmds)

        other_cmds = flat

        # ---------------------------------------------
        # Keyword detection (same as exception.bat)
        # ---------------------------------------------
        text = "\n".join(lines).lower()

        def flag(*keywords):
            return "yes" if any(k in text for k in keywords) else ""

        def sb_flag():
            import re
            if re.search(r"\bsb(?!laster)\b", text):
                return "yes"
            if "soundblaster" in text or "sound blaster" in text:
                return "yes"
            if "adlib" in text:
                return "yes"
            if "opl" in text:
                return "yes"
            return ""

        # IBM only when used as a selectable option
        ibm_flag = ""
        if "ibm" in text:
            if any(v in text for v in ("cga", "ega", "vga", "mcga", "tandy", "pcjr", "machine")):
                ibm_flag = "yes"

        # ---------------------------------------------
        # Return run_* fields (exception_* minus 3)
        # ---------------------------------------------
        return {
            "exists": "yes",
            "notes": "; ".join(notes),
            "other": "; ".join(other_cmds),
            "mixer": "; ".join(mixer_cmds),        
            "autotype": "; ".join(autotype_cmds),  
            "config": "; ".join(config_cmds),

            # keyword flags identical to exception.bat
            "hercules": flag("hercules"),
            "cga": flag("cga"),
            "composite": flag("composite"),
            "ega": flag("ega"),
            "vga": flag("vga"),
            "tandy": flag("tandy"),
            "pcjr": flag("pcjr"),
            "ibm": ibm_flag,
            "pcspeaker": flag("pc speaker", "pcspeaker", "pc-speaker"),
            "ps1": flag("ps1", "ps\\1"),
            "innovation": flag("innovation"),
            "imfc": flag("imfc", "music feature"),
            "covox_disney": flag("covox", "disney"),
            "cms_gb": flag("cms", "game blaster", "gameblaster", "game-blaster", "gb "),
            "sb_adlib_opl": sb_flag(),
            "roland_midi_gm": flag("roland", "midi", "general midi"),
            "mt32": flag("mt32", "mt-32", "mt 32", "lapc"),
            "sc55": flag("soundcanvas", "sound canvas", "sound-canvas", "ssc", "sc55", "sc-55", "fluidsynth"),
            
            "network": "yes" if ("network" in text or "multiplayer" in text) else "",
        }

    # -----------------------------------------------------
    # 5. Follow run → runH → run2 chain
    # -----------------------------------------------------
    chain = []
    visited = set()

    current_key = start_file

    with zipfile.ZipFile(zip_path, "r") as z:
        while current_key and current_key not in visited:
            visited.add(current_key)

            raw = z.read(current_key).decode("utf-8", errors="ignore").splitlines()
            parsed = parse_single_run(raw)
            chain.append(parsed)

            # Detect next run file
            next_key = None
            for line in raw:
                stripped = line.strip()
                parts = stripped.split()
                if not parts:
                    continue

                token = parts[0].lower()

                # direct run
                if token.startswith("run"):
                    stem = Path(token).stem.lower()
                    # find matching file
                    for fullpath in run_files:
                        if Path(fullpath).stem.lower() == stem:
                            next_key = run_files[fullpath]
                            break
                    break

                # call run
                if token == "call" and len(parts) > 1:
                    stem = Path(parts[1]).stem.lower()
                    for fullpath in run_files:
                        if Path(fullpath).stem.lower() == stem:
                            next_key = run_files[fullpath]
                            break
                    break

            current_key = next_key

    return chain

# ---------------------------------------------------------
# MAIN SCAN
# ---------------------------------------------------------
def main():
    global ticker_running, ticker_paused
    forks = load_dosbox_forks(DOSBOX_TXT)
    failed_folders = []

    headers = [
        "Full Name", "Project", "Folder Name", "Year", "Dosbox Fork",
        "conf_file name", "conf_fork_label",

        # Conf Settings
        "conf_output", "conf_fullresolution", "conf_Forced_Aspect", "conf_machine", "conf_composite_CGA",
        "conf_core", "conf_cputype", "conf_cycles/cpu_cycles", "conf_cpu_throttle", "conf_memsize", "conf_vmemsize",
        "conf_vmem_delay", "conf_dos_rate", "conf_vesa_modes", "conf_vga_render_per_scanline",
        "conf_mouse_sensitivity",

        # Sound
        "conf_mididevice", "conf_mpu401", "conf_mt32_romdir", "conf_sbtype", "conf_sbmixer",
        "conf_oplmode", "conf_cms", "conf_oplemu", "conf_tandy", "conf_gus", "conf_pcspeaker",
        "conf_disney/lpt_dac", "conf_innovation", "conf_imfc", "conf_PS1",

        # Other
        "conf_reelmagic", "conf_joystick_type", "conf_serial1", "conf_xms", "conf_ems",
        "conf_umb", "conf_pcjr_memory_config", "conf_file_locking", "conf_ipx", "conf_printer",

        # Autoexec
        "conf_autoexec_summary", "conf_internal_path", "conf_boot", "conf_executed_program", "conf_autoexec_other", "conf_set",
        
        # Exception.bat
        "exception_exists", "exception_notes", "exception_86box",
        "exception_scummvm", "exception_hercules", "exception_cga", "exception_composite",
        "exception_ega", "exception_vga", "exception_tandy", "exception_pcjr",
        "exception_ibm", "exception_pcspeaker", "exception_ps1", "exception_innovation",
        "exception_imfc", "exception_covox_disney", "exception_cms_gb", "exception_sb_adlib_opl",
        "exception_roland_midi_gm", "exception_mt32", "exception_sc55", "exception_external", "exception_other",
        
        # Run.bat dynamic columns will be appended later
    ]

    rows = []

    max_run_depth = 0

    # Ask user how many folders to scan
    limit = ask_folder_limit()
    
    # Start ticker AFTER user input
    t = threading.Thread(target=ticker, daemon=True)
    t.start()

    # Define projects
    projects = [
        {
            "label": "IBM PC",
            "base": EXODOS_BASE,
            "dos_root": EXODOS_BASE / "!dos",
            "platform_xml": "MS-DOS.xml",
        },
        {
            "label": "IBM PCjr",
            "base": PCJR_BASE,
            "dos_root": PCJR_BASE / "!pcjr",
            "platform_xml": "PCjr.xml",
        },
    ]

    # Preload XML mappings per project
    xml_maps = {
        p["label"]: load_launchbox_playlists(LAUNCHBOX_ROOT, p["platform_xml"])
        for p in projects
    }

    for project in projects:
        project_label = project["label"]
        project_base  = project["base"]
        dos_root      = project["dos_root"]
        xml_playlists_map = xml_maps[project_label]

        if not dos_root.exists():
            print(f"WARNING: DOS root for {project_label} not found at {dos_root}")
            continue

        folders = [f for f in sorted(dos_root.iterdir()) if f.is_dir()]

        if limit is not None:
            folders = folders[:limit]

        for game_folder in folders:
            try:
                folder_name = game_folder.name

                # Launcher
                full_name, year, launcher_bat = find_launcher(game_folder)
                launcher_key = launcher_bat.name.strip().strip('"').strip("'").lower().replace("\\", "/") if launcher_bat else ""
                if launcher_key.endswith(".bat"):
                    launcher_key = launcher_key[:-4]
                fork = forks.get(launcher_key, "MISSING")
                
                zip_name = launcher_bat.name.replace(".bat", ".zip") if launcher_bat else ""
                
                # Exception.bat
                exception_data = parse_exception_bat(game_folder)

                # Conf files
                for conf in game_folder.glob("*.conf"):
                    conf_label = read_first_line(conf)
                    parsed = parse_conf(conf)

                    # ---------------------------------------------
                    # Decide if this conf actually launches run*.bat
                    # ---------------------------------------------
                    raw_cmd = parsed["real_final_cmd"].strip()

                    # Remove leading @
                    if raw_cmd.startswith("@"):
                        raw_cmd = raw_cmd[1:].lstrip()

                    parts = raw_cmd.split()

                    if not parts:
                        exec_stem = ""
                    else:
                        # Handle "call runX"
                        if parts[0].lower() == "call" and len(parts) > 1:
                            token = parts[1]
                        else:
                            token = parts[0]

                        exec_stem = Path(token).stem.lower()

                    # ---------------------------------------------
                    # Only parse run-chain if it really is run*.bat
                    # ---------------------------------------------
                    if exec_stem.startswith("run"):
                        run_chain = parse_run_chain(
                            game_folder,
                            parsed["real_final_cmd"],
                            parsed["internal_path"],
                            zip_name,
                            project_base,
                        )
                    else:
                        run_chain = []

                    max_run_depth = max(max_run_depth, len(run_chain))


                    # Build row
                    row = [
                        full_name or "",
                        project_label,
                        folder_name,
                        year or "",
                        fork,

                        conf.name,
                        conf_label,

                        parsed["output"],
                        parsed["fullresolution"],
                        parsed["forced_aspect"],
                        parsed["machine"],
                        parsed["composite"],
                        parsed["core"],
                        parsed["cputype"],
                        parsed["cycles"],
                        parsed["cpu_throttle"],
                        parsed["memsize"],
                        parsed["vmemsize"],
                        parsed["vmem_delay"],
                        parsed["dos_rate"],
                        parsed["vesa_modes"],
                        parsed["vga_render_per_scanline"],
                        parsed["mouse_sensitivity"],

                        parsed["mididevice"],
                        parsed["mpu401"],
                        parsed["mt32_romdir"],
                        parsed["sbtype"],
                        parsed["sbmixer"],
                        parsed["oplmode"],
                        parsed["cms"],
                        parsed["oplemu"],
                        parsed["tandy"],
                        parsed["gus"],
                        parsed["pcspeaker"],
                        parsed["disney"],
                        parsed["innovation"],
                        parsed["imfc"],
                        parsed["ps1"],

                        parsed["reelmagic"],
                        parsed["joysticktype"],
                        parsed["serial"],
                        parsed["xms"],
                        parsed["ems"],
                        parsed["umb"],
                        parsed["pcjr_memory_config"],
                        parsed["file_locking"],
                        parsed["ipx"],
                        parsed["printer"],

                        parsed["autoexec_summary"],
                        parsed["internal_path"],
                        parsed["boot_cmd"],
                        parsed["real_final_cmd"],
                        parsed["other_autoexec"],
                        parsed["set_cmds"],
                        
                        exception_data["exception_exists"],
                        exception_data["exception_notes"],
                        exception_data["exception_86box"],
                        exception_data["exception_scummvm"],
                        exception_data["exception_hercules"],
                        exception_data["exception_cga"],
                        exception_data["exception_composite"],
                        exception_data["exception_ega"],
                        exception_data["exception_vga"],
                        exception_data["exception_tandy"],
                        exception_data["exception_pcjr"],
                        exception_data["exception_ibm"],
                        exception_data["exception_pcspeaker"],
                        exception_data["exception_ps1"],
                        exception_data["exception_innovation"],
                        exception_data["exception_imfc"],
                        exception_data["exception_covox_disney"],
                        exception_data["exception_cms_gb"],
                        exception_data["exception_sb_adlib_opl"],
                        exception_data["exception_roland_midi_gm"],
                        exception_data["exception_mt32"],
                        exception_data["exception_sc55"],
                        exception_data["exception_external"],
                        exception_data["exception_other"],
                    ]

                    # Add run chain columns for this row
                    for i in range(1, max_run_depth + 1):
                        if i <= len(run_chain):
                            r = run_chain[i-1]
                            row.extend([
                                r["exists"], r["notes"], r["other"], r["mixer"], r["autotype"], r["config"],
                                r["hercules"], r["cga"], r["composite"], r["ega"], r["vga"],
                                r["tandy"], r["pcjr"], r["ibm"], r["pcspeaker"], r["ps1"],
                                r["innovation"], r["imfc"], r["covox_disney"], r["cms_gb"],
                                r["sb_adlib_opl"], r["roland_midi_gm"], r["mt32"], r["sc55"], r["network"],
                            ])
                        else:
                            row.extend([""] * 25)
                    
                    lookup_key = launcher_bat.stem if launcher_bat else (full_name or "")
                    xml_playlists = xml_playlists_map.get(lookup_key, "")

                    row.append(xml_playlists)

                    rows.append(row)

            except Exception as e:
                print(f"\nERROR processing folder: {game_folder}")
                print(f"Reason: {e}")
                print("Continuing to next folder...\n")
                failed_folders.append((game_folder, str(e)))
                continue

    # ---------------------------------------------------------
    # Add dynamic run columns BEFORE writing header row
    # ---------------------------------------------------------
    for i in range(1, max_run_depth + 1):
        prefix = f"run{i}_"
        headers.extend([
            prefix + "exists",
            prefix + "notes",
            prefix + "other",
            prefix + "mixer",
            prefix + "autotype",
            prefix + "config",
            prefix + "hercules",
            prefix + "cga",
            prefix + "composite",
            prefix + "ega",
            prefix + "vga",
            prefix + "tandy",
            prefix + "pcjr",
            prefix + "ibm",
            prefix + "pcspeaker",
            prefix + "ps1",
            prefix + "innovation",
            prefix + "imfc",
            prefix + "covox_disney",
            prefix + "cms_gb",
            prefix + "sb_adlib_opl",
            prefix + "roland_midi_gm",
            prefix + "mt32",
            prefix + "sc55",
            prefix + "network",
        ])
    headers.append("xml_playlists")
    
    # ---------------------------------------------------------
    # Sort final output by Full Name
    # ---------------------------------------------------------
    rows.sort(key=lambda r: r[0].lower())

    # ---------------------------------------------------------
    # Write CSV
    # ---------------------------------------------------------
    try:
        with OUTPUT_CSV.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for row in rows:
                writer.writerow([safe_csv_value(v) for v in row])
    except PermissionError:
        print("\nERROR: Cannot write to the output CSV file.")
        print(f"The file is currently open:\n{OUTPUT_CSV}")
        print("Please close it and run the scan again.")
        ticker_paused = True
        input("\nPress Enter to exit...")
        ticker_paused = False
        return

    # ---------------------------------------------------------
    # Write XLSX (formatted)
    # ---------------------------------------------------------
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as e:
        print("\nopenpyxl is not installed, skipping XLSX generation.")
        print("Error:", e)
        ticker_paused = True
        input("\nPress Enter to exit...")
        ticker_paused = False
        return


    def write_xlsx(csv_path, xlsx_path):
        import csv
        import re

        ILLEGAL_CHARACTERS_RE = re.compile(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]')

        def clean_excel_value(value):
            if isinstance(value, str):
                return ILLEGAL_CHARACTERS_RE.sub("", value)
            return value

        wb = Workbook()
        ws = wb.active
        ws.title = "Scan Output"

        # Read CSV
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            for r, row in enumerate(reader, start=1):
                game_name = row[0] if len(row) > 0 else ""
                folder_name = row[1] if len(row) > 1 else ""

                for c, value in enumerate(row, start=1):
                    if isinstance(value, str) and ILLEGAL_CHARACTERS_RE.search(value):
                        print(
                            f"Cleaned illegal characters in "
                            f"[{game_name}] ({folder_name}), row {r}, column {c}: {repr(value)}"
                        )

                    ws.cell(row=r, column=c, value=clean_excel_value(value))
                    
        # Add word wrap to all cells
        from openpyxl.styles import Alignment
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrapText=True)
                
        # Auto-adjust row height
        for row in ws.row_dimensions.values():
            row.height = None
        
        # Shading
        from openpyxl.styles import PatternFill

        # Define two alternating fill colours
        fill_a = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")  # light blue
        fill_b = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")  # white

        # Columns that define a unique game identity
        group_columns = [1, 2, 3, 4, 5]  # Full Name, Project, Folder Name, Year, Dosbox Fork

        current_fill = fill_a
        start_row = 2  # skip header

        while start_row <= ws.max_row:
            end_row = start_row
            # Find consecutive rows belonging to the same game
            while end_row + 1 <= ws.max_row:
                same_game = all(
                    ws.cell(row=start_row, column=c).value == ws.cell(row=end_row + 1, column=c).value
                    for c in group_columns
                )
                if same_game:
                    end_row += 1
                else:
                    break

            # Apply shading to all rows in this group
            for r in range(start_row, end_row + 1):
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = current_fill

            # Alternate colour for next group
            current_fill = fill_b if current_fill == fill_a else fill_a
            start_row = end_row + 1
        
        # white border
        from openpyxl.styles import Border, Side

        thin_white = Side(style="thin", color="FFFFFF")

        # Apply thin white border to all cells
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).border = Border(
                    left=thin_white,
                    right=thin_white,
                    top=thin_white,
                    bottom=thin_white
                )
            
        # Format header row and adjust column widths
        from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_border = Border(bottom=Side(style="medium", color="FFFFFF"))
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", wrapText=True)

            # Calculate max content length in this column
            max_len = max(
                (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)),
                default=10
            )

            # Add padding for filter arrow and readability
            adjusted_width = max_len + 10  # +10 gives room for dropdown arrow
            ws.column_dimensions[get_column_letter(col)].width = min(adjusted_width, 60)
            
        # Build a lookup of header → column index
        header_to_col = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        
        # Find the last run_* column dynamically
        last_run_col = None
        for header, col in header_to_col.items():
            if header.startswith("run") and header.endswith("_network"):
                if last_run_col is None or col > last_run_col:
                    last_run_col = col

        # Section boundaries (vertical borders)
        section_splits = [
            header_to_col["Dosbox Fork"],
            header_to_col["conf_set"],
            header_to_col["exception_other"],
            last_run_col,  # divider BEFORE xml_playlists
        ]
        
        from openpyxl.styles import Border, Side

        thin_black = Side(style="thin", color="000000")

        for split_col in section_splits:
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=split_col)
                existing = cell.border

                cell.border = Border(
                    left=existing.left,
                    top=existing.top,
                    bottom=existing.bottom,
                    right=thin_black
                )

        ws.freeze_panes = "F2"
        ws.auto_filter.ref = ws.dimensions

        wb.save(xlsx_path)


    import traceback

    try:
        xlsx_path = OUTPUT_CSV.with_suffix(".xlsx")
        write_xlsx(OUTPUT_CSV, xlsx_path)
        print(f"Formatted XLSX written to:\n{xlsx_path}")
    except PermissionError:
        print(f"\nERROR: Cannot write to {xlsx_path}")
        print("It looks like the file is open in Excel.")
        print("Please close it, then press Enter to retry writing the XLSX.")
        ticker_paused = True
        input("> ")
        ticker_paused = False
        try:
            write_xlsx(OUTPUT_CSV, xlsx_path)
            print(f"\nFormatted XLSX written successfully after retry:\n{xlsx_path}")
        except PermissionError:
            print("\nStill cannot write to the XLSX file. Please ensure it’s closed and rerun the script.")
    except Exception as e:
        print("\nERROR while writing XLSX:")
        print(e)
        import traceback
        traceback.print_exc()
        ticker_paused = True
        input("\nPress Enter to exit...")
        ticker_paused = False
        return

    # ---------------------------------------------------------
    # Error summary
    # ---------------------------------------------------------
    if failed_folders:
        print("\nThe following folders failed during scanning:")
        for folder, reason in failed_folders:
            print(f" - {folder}: {reason}")
    else:
        print("\nNo folder errors detected.")
    
    ticker_running = False
    time.sleep(0.2)  # allow final frame to print cleanly
    sys.stdout.write("\n")  # move to a new line so ticker and prompt don't overlap
    sys.stdout.flush()
    
    input("\nPress Enter to exit...")

if __name__ == "__main__":
    main()